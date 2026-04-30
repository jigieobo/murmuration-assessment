"""Murmuration Technical Analyst Assessment - Part 2 MVP
========================================================

A single-file pipeline that implements the core architecture from the
Part 1 design doc, scoped to Rhode Island municipal elected officials.

"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
import sys
import tempfile
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook
from rapidfuzz import fuzz, process


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

STATE_FIPS = "44"  # Rhode Island
STATE_ABBR = "RI"
STATE_NAME = "Rhode Island"

DEFAULT_DB_PATH = "ri_officials.db"
RAW_DIR = Path("data/raw")

# Reference data (Tier 1): U.S. Census Bureau
# 2020 Decennial Census PL94-171 endpoint - small, public, no key needed.
CENSUS_COUNTY_URL = (
    f"https://api.census.gov/data/2020/dec/pl"
    f"?get=NAME&for=county:*&in=state:{STATE_FIPS}"
)
CENSUS_COUSUB_URL = (
    f"https://api.census.gov/data/2020/dec/pl"
    f"?get=NAME,P1_001N&for=county%20subdivision:*&in=state:{STATE_FIPS}"
)

# Tier 1: officials (RI SOS)
# RI SOS publishes per-municipality JSON files. The human-facing landing page
# lives on www.ri.gov; the actual JSON files are served from S3. The script
# loops over the seeded municipalities and constructs:
#     {RI_SOS_BASE_FETCH_URL}/{slug}.json
# where `slug` is the muni name lowercased with spaces replaced by '_' (the
# convention barrington.json suggests; verified against multi-word munis only
# at run time, with per-slug failures logged and skipped).
TIER1_NAME = "Rhode Island Secretary of State - Elections Division"
TIER1_URL  = "https://www.ri.gov/election/results/2024/general_election/"  # human-facing index
RI_SOS_BASE_FETCH_URL = (
    "https://rigov.s3.amazonaws.com/election/results/2024/general_election"
)
TIER1_SLUG = "ri_sos"

# Tier 2: cross-check (data.ri.gov)
# The election-summary XLSX (Gen24EX.xlsx for the 2024 General Election)
# is bundled with the script. The portal updates this file per cycle;
# swap the bundled file for a different cycle. Path is resolved relative
# to this script so `python ri_officials_mvp.py run` works from any CWD.
TIER2_NAME = "Rhode Island Open Data Portal"
TIER2_URL  = "https://data.ri.gov/"
TIER2_FILENAME = "Gen24EX.xlsx"
TIER2_PATH = Path(__file__).resolve().parent / TIER2_FILENAME

CENSUS_NAME = "U.S. Census Bureau"
CENSUS_URL  = "https://api.census.gov/"

USER_AGENT   = "murmuration-eo-mvp/0.3 (technical-assessment)"
HTTP_TIMEOUT = 30

# Fuzzy thresholds. Tuneable.
# - Title matching uses token_set_ratio (word-set similarity) and is lenient
#   because we're comparing a raw title to a curated synonym phrase.
# - Name matching uses ratio (Levenshtein-ish) and is conservative because
#   a wrong merge of two officials is much worse than a duplicate row.
FUZZY_TITLE_THRESHOLD = 80
FUZZY_NAME_THRESHOLD  = 90


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------
# Retains the Part 1 deviation (parallel `municipalities` table; offices
# points to county XOR muni). Adds two tables that earn their keep at MVP
# scale: `office_type_synonyms` for fuzzy reconciliation, `needs_review`
# for failed-record provenance, and `collection_log` for fetch audit.
# Full schema (incl. contact_info, addresses, indexes) is in
# reference_implementation/schema.sql.

SCHEMA_SQL = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS states (
    state_fips  TEXT PRIMARY KEY,
    state_abbr  TEXT NOT NULL UNIQUE,
    state_name  TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS counties (
    county_fips             TEXT PRIMARY KEY,
    state_fips              TEXT NOT NULL REFERENCES states(state_fips),
    county_name             TEXT NOT NULL,
    county_type             TEXT NOT NULL DEFAULT 'county' CHECK (county_type IN (
                                'county','parish','borough','independent_city','census_area'
                            )),
    government_form         TEXT NOT NULL CHECK (government_form IN (
                                'standard','consolidated_city_county','no_county_government'
                            )),
    has_elected_government  INTEGER NOT NULL CHECK (has_elected_government IN (0,1)),
    official_website        TEXT,
    population              INTEGER,
    last_verified_at        TEXT
);

-- DEVIATION FROM PART 1: parallel municipalities table.
CREATE TABLE IF NOT EXISTS municipalities (
    muni_id           INTEGER PRIMARY KEY AUTOINCREMENT,
    state_fips        TEXT NOT NULL REFERENCES states(state_fips),
    county_fips       TEXT REFERENCES counties(county_fips),
    muni_name         TEXT NOT NULL,
    muni_type         TEXT NOT NULL,
    population        INTEGER,
    official_website  TEXT,
    last_verified_at  TEXT,
    UNIQUE (state_fips, muni_name)
);

CREATE TABLE IF NOT EXISTS office_types (
    office_type_id      INTEGER PRIMARY KEY AUTOINCREMENT,
    canonical_name      TEXT NOT NULL UNIQUE,
    description         TEXT,
    is_typically_board  INTEGER NOT NULL DEFAULT 0 CHECK (is_typically_board IN (0,1))
);

CREATE TABLE IF NOT EXISTS office_type_synonyms (
    synonym_id     INTEGER PRIMARY KEY AUTOINCREMENT,
    office_type_id INTEGER NOT NULL REFERENCES office_types(office_type_id),
    synonym        TEXT NOT NULL,
    UNIQUE (office_type_id, synonym)
);

CREATE TABLE IF NOT EXISTS offices (
    office_id          INTEGER PRIMARY KEY AUTOINCREMENT,
    county_fips        TEXT REFERENCES counties(county_fips),
    muni_id            INTEGER REFERENCES municipalities(muni_id),
    office_type_id     INTEGER NOT NULL REFERENCES office_types(office_type_id),
    local_title        TEXT NOT NULL,
    district_or_seat   TEXT,
    is_partisan        INTEGER CHECK (is_partisan IN (0,1)),
    term_length_years  INTEGER,
    notes              TEXT,
    CHECK (
        (county_fips IS NOT NULL AND muni_id IS NULL)
     OR (county_fips IS NULL     AND muni_id IS NOT NULL)
    )
);

CREATE TABLE IF NOT EXISTS officials (
    official_id  INTEGER PRIMARY KEY AUTOINCREMENT,
    full_name    TEXT NOT NULL,
    first_name   TEXT,
    last_name    TEXT,
    party        TEXT,
    notes        TEXT
);

CREATE TABLE IF NOT EXISTS office_holders (
    holder_id          INTEGER PRIMARY KEY AUTOINCREMENT,
    official_id        INTEGER NOT NULL REFERENCES officials(official_id),
    office_id          INTEGER NOT NULL REFERENCES offices(office_id),
    term_start         TEXT,
    term_end           TEXT,
    is_current         INTEGER NOT NULL DEFAULT 1 CHECK (is_current IN (0,1)),
    assumption_method  TEXT CHECK (assumption_method IN (
                            'elected','appointed','interim','acting'
                        )),
    source_id          INTEGER REFERENCES sources(source_id),
    collected_at       TEXT,
    last_verified_at   TEXT
);

-- contact_info and addresses are part of the Part 1 schema.
-- The MVP creates the tables but does NOT populate them: RI SOS election
-- results don't carry contact data. Filling them needs a secondary source
-- (county/muni directory scrape, RILA association roster, etc.) - tracked
-- in the with-more-time list of the README.
CREATE TABLE IF NOT EXISTS contact_info (
    contact_id      INTEGER PRIMARY KEY AUTOINCREMENT,
    official_id     INTEGER REFERENCES officials(official_id),
    office_id       INTEGER REFERENCES offices(office_id),
    contact_type    TEXT NOT NULL CHECK (contact_type IN ('phone','fax','email','website')),
    location_label  TEXT,
    value           TEXT NOT NULL,
    CHECK (
        (official_id IS NOT NULL AND office_id IS NULL)
     OR (official_id IS NULL     AND office_id IS NOT NULL)
    )
);

CREATE TABLE IF NOT EXISTS addresses (
    address_id      INTEGER PRIMARY KEY AUTOINCREMENT,
    official_id     INTEGER REFERENCES officials(official_id),
    office_id       INTEGER REFERENCES offices(office_id),
    location_label  TEXT,
    street          TEXT,
    city            TEXT,
    state           TEXT,
    zip             TEXT,
    CHECK (
        (official_id IS NOT NULL AND office_id IS NULL)
     OR (official_id IS NULL     AND office_id IS NOT NULL)
    )
);

CREATE TABLE IF NOT EXISTS sources (
    source_id          INTEGER PRIMARY KEY AUTOINCREMENT,
    source_name        TEXT NOT NULL UNIQUE,
    source_url         TEXT,
    source_type        TEXT NOT NULL CHECK (source_type IN (
                            'official_website','state_portal','ballotpedia',
                            'manual','api','open_data_portal','aggregator'
                       )),
    reliability_tier   INTEGER NOT NULL CHECK (reliability_tier BETWEEN 1 AND 4),
    last_fetched_at    TEXT,
    notes              TEXT
);

CREATE TABLE IF NOT EXISTS collection_log (
    log_id             INTEGER PRIMARY KEY AUTOINCREMENT,
    county_fips        TEXT REFERENCES counties(county_fips),
    muni_id            INTEGER REFERENCES municipalities(muni_id),
    source_id          INTEGER REFERENCES sources(source_id),
    run_at             TEXT NOT NULL,
    status             TEXT NOT NULL,
    records_found      INTEGER,
    records_updated    INTEGER,
    storage_path       TEXT,
    raw_content_hash   TEXT,
    error_message      TEXT
);

CREATE TABLE IF NOT EXISTS needs_review (
    review_id      INTEGER PRIMARY KEY AUTOINCREMENT,
    raw_record     TEXT NOT NULL,
    failure_reason TEXT NOT NULL,
    source_id      INTEGER REFERENCES sources(source_id),
    storage_path   TEXT,
    flagged_at     TEXT NOT NULL
);
"""


# ---------------------------------------------------------------------------
# Static catalog: office types + their synonyms
# ---------------------------------------------------------------------------
# Each canonical office_type carries an inline list of synonym phrases.
# At seed time the canonical row goes into `office_types` and each
# synonym becomes a row in `office_type_synonyms`. The reconciler then
# fuzzy-matches raw titles against the union of all synonyms.
#
# This is the table that does cross-state title normalization. RI's
# "Town Council Member" and NC's "County Commissioner" both reconcile
# to canonical_name = "Local Legislative Member" because both phrases
# appear under the same canonical's synonym list.

OFFICE_TYPES_WITH_SYNONYMS: list[dict] = [
    {
        "canonical_name": "Chief Local Executive",
        "description": (
            "Single elected head of a local-government unit. Generalizes "
            "Mayor (city/town) and County Executive."
        ),
        "is_typically_board": False,
        "synonyms": [
            "mayor", "city mayor", "town mayor",
            "county executive", "elected town manager",
        ],
    },
    {
        "canonical_name": "Local Legislative Member",
        "description": (
            "Member of a multi-seat elected legislative body for the "
            "operative unit of local government."
        ),
        "is_typically_board": True,
        "synonyms": [
            "town council", "city council", "town council member",
            "city council member", "council member", "councilor",
            "councilman", "councilwoman",
            "alderman", "alderwoman", "alderperson",
            "selectman", "selectwoman", "selectperson",
            "county commissioner", "county supervisor",
            "freeholder", "board of supervisors",
        ],
    },
    {
        "canonical_name": "Local Legislative Chair",
        "description": (
            "Distinctly-elected presiding officer of a local "
            "legislative body."
        ),
        "is_typically_board": False,
        "synonyms": [
            "town council president", "city council president",
            "council president", "board chair", "commissioners chair",
        ],
    },
    {
        "canonical_name": "School Committee Member",
        "description": (
            "Member of a multi-seat elected board overseeing local "
            "public education."
        ),
        "is_typically_board": True,
        "synonyms": [
            "school committee", "school committee member",
            "school board", "school board member",
            "board of education", "board of education member",
        ],
    },
    {
        "canonical_name": "Local Records Clerk",
        "description": (
            "Elected keeper of vital records for the local-government "
            "unit. Excludes appointed clerks."
        ),
        "is_typically_board": False,
        "synonyms": [
            "town clerk", "city clerk", "county clerk",
            "register of deeds", "recorder of deeds",
        ],
    },
    {
        "canonical_name": "Town Moderator",
        "description": (
            "Elected presiding officer of the town meeting form of "
            "government. RI/MA-specific."
        ),
        "is_typically_board": False,
        "synonyms": ["town moderator", "moderator"],
    },
]

# (source_name, source_url, source_type, reliability_tier)
SOURCES: list[tuple[str, str, str, int]] = [
    (TIER1_NAME,  TIER1_URL,  "state_portal",     1),
    (CENSUS_NAME, CENSUS_URL, "api",              1),
    (TIER2_NAME,  TIER2_URL,  "open_data_portal", 2),
]


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def open_db(db_path: str | Path) -> sqlite3.Connection:
    """Open a SQLite connection with foreign-key enforcement turned on
    (off by default in SQLite) and Row access for cleaner column lookup."""
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row
    return conn


def init_schema(conn: sqlite3.Connection) -> None:
    """Apply SCHEMA_SQL. Idempotent - all CREATE statements use IF NOT EXISTS."""
    conn.executescript(SCHEMA_SQL)
    conn.commit()


def seed_static(conn: sqlite3.Connection) -> dict[str, int]:
    """Insert the data we know without fetching: state + sources +
    canonical office_types + synonyms.

    `office_type_synonyms` is populated in a loop, one synonym per row,
    so the reconciler can join straight against it. The hardcoded
    OFFICE_TYPES_WITH_SYNONYMS list is the bootstrap prior - the
    aligned v2 plan grows this table from analyst-confirmed
    needs_review entries (see Section 5 of the design doc).
    """
    conn.execute(
        "INSERT OR IGNORE INTO states (state_fips, state_abbr, state_name) "
        "VALUES (?, ?, ?)",
        (STATE_FIPS, STATE_ABBR, STATE_NAME),
    )
    conn.executemany(
        "INSERT OR IGNORE INTO sources "
        "(source_name, source_url, source_type, reliability_tier) "
        "VALUES (?, ?, ?, ?)",
        SOURCES,
    )

    n_synonyms = 0
    for entry in OFFICE_TYPES_WITH_SYNONYMS:
        conn.execute(
            "INSERT OR IGNORE INTO office_types "
            "(canonical_name, description, is_typically_board) "
            "VALUES (?, ?, ?)",
            (entry["canonical_name"], entry["description"],
             1 if entry["is_typically_board"] else 0),
        )
        ot_id = conn.execute(
            "SELECT office_type_id FROM office_types WHERE canonical_name = ?",
            (entry["canonical_name"],),
        ).fetchone()[0]
        for synonym in entry["synonyms"]:
            conn.execute(
                "INSERT OR IGNORE INTO office_type_synonyms "
                "(office_type_id, synonym) VALUES (?, ?)",
                (ot_id, synonym),
            )
            n_synonyms += 1
    conn.commit()
    return {
        "states":       conn.execute("SELECT COUNT(*) FROM states").fetchone()[0],
        "sources":      conn.execute("SELECT COUNT(*) FROM sources").fetchone()[0],
        "office_types": conn.execute("SELECT COUNT(*) FROM office_types").fetchone()[0],
        "synonyms":     conn.execute("SELECT COUNT(*) FROM office_type_synonyms").fetchone()[0],
    }


# ---------------------------------------------------------------------------
# Reference data fetch (Census)
# ---------------------------------------------------------------------------

def _http_get_json(url: str) -> object:
    """GET a URL, return parsed JSON. Hard-fails with a clear message."""
    try:
        resp = requests.get(
            url, headers={"User-Agent": USER_AGENT}, timeout=HTTP_TIMEOUT
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            f"Census API fetch failed for {url}\n"
            f"  reason: {type(exc).__name__}: {exc}\n"
            "  Workaround (with more time, would be more graceful): rerun with\n"
            "  --reference-from-file pointing at a saved Census response."
        ) from exc


def _parse_census_table(payload) -> list[dict]:
    """Census API returns [[headers], [row], [row], ...].

    Convert into a list of dicts so downstream loops can index by name.
    """
    if not isinstance(payload, list) or len(payload) < 2:
        raise ValueError(f"unexpected Census payload shape: {type(payload).__name__}")
    headers = payload[0]
    return [dict(zip(headers, row)) for row in payload[1:]]


# Census county subdivision NAMEs look like:
#   "Providence city, Providence County, Rhode Island"
#   "Bristol town, Bristol County, Rhode Island"
# The first comma-segment is "<muni_name> <muni_type>".
_COUSUB_NAME_RE = re.compile(
    r"^(?P<name>.+?)\s+(?P<type>city|town|township|borough|village)$",
    re.IGNORECASE,
)


def _split_cousub_name(name_field: str) -> tuple[str, str, str]:
    """Return (muni_name, muni_type, county_name) from a Census NAME."""
    parts = [p.strip() for p in name_field.split(",")]
    if len(parts) < 2:
        raise ValueError(f"unexpected NAME field shape: {name_field!r}")
    place_part, county_part = parts[0], parts[1]
    m = _COUSUB_NAME_RE.match(place_part)
    if not m:
        # Fallback for non-standard suffixes; default type is 'town'.
        return place_part, "town", county_part
    return m.group("name").strip(), m.group("type").lower(), county_part


def fetch_census_counties(*, from_file: str | Path | None = None) -> list[dict]:
    """Pull all counties for STATE_FIPS from Census 2020 PL data.

    Returns a list of dicts:
        {county_fips, state_fips, county_name}

    Loops the API response into the desired data structure.
    """
    if from_file:
        payload = json.loads(Path(from_file).read_text(encoding="utf-8"))
    else:
        payload = _http_get_json(CENSUS_COUNTY_URL)

    out: list[dict] = []
    for row in _parse_census_table(payload):
        # NAME = "Bristol County, Rhode Island"
        county_name = row["NAME"].split(",", 1)[0].strip()
        county_fips = f"{row['state']}{row['county']}"
        out.append({
            "county_fips": county_fips,
            "state_fips":  row["state"],
            "county_name": county_name,
        })
    return out


def fetch_census_municipalities(*, from_file: str | Path | None = None) -> list[dict]:
    """Pull RI cities/towns from Census County Subdivisions.

    For RI, county subdivisions == municipalities exactly. The endpoint
    returns one row per cousub with population (P1_001N) included.

    Returns:
        [{state_fips, county_name, county_fips, muni_name, muni_type, population}, ...]
    """
    if from_file:
        payload = json.loads(Path(from_file).read_text(encoding="utf-8"))
    else:
        payload = _http_get_json(CENSUS_COUSUB_URL)

    out: list[dict] = []
    for row in _parse_census_table(payload):
        muni_name, muni_type, county_name = _split_cousub_name(row["NAME"])
        # RI county subdivisions occasionally include a "Block Island" or
        # similar parenthetical; defensive: require name be non-empty.
        if not muni_name:
            continue
        out.append({
            "state_fips":  row["state"],
            "county_fips": f"{row['state']}{row['county']}",
            "county_name": county_name,
            "muni_name":   muni_name,
            "muni_type":   muni_type,
            "population":  int(row["P1_001N"]) if row.get("P1_001N") else None,
        })
    return out


def seed_reference_from_census(
    conn: sqlite3.Connection,
    *,
    counties_from_file: str | Path | None = None,
    munis_from_file:    str | Path | None = None,
) -> dict[str, int]:
    """Pull RI counties + municipalities from Census, loop into the DB.

    For RI, every county is loaded with `government_form =
    'no_county_government'` and `has_elected_government = 0`. This is
    correct for RI and what makes the no-elected-county-gov assertion
    visible in the data.
    """
    counties = fetch_census_counties(from_file=counties_from_file)
    n_counties = 0
    for c in counties:
        # county_type defaults to 'county' for RI - the schema's enum also
        # supports parish/borough/independent_city/census_area for states
        # where Census uses different terminology.
        conn.execute(
            """
            INSERT OR IGNORE INTO counties
                (county_fips, state_fips, county_name, county_type,
                 government_form, has_elected_government,
                 official_website, population, last_verified_at)
            VALUES (?, ?, ?, 'county', ?, ?, NULL, NULL, ?)
            """,
            (c["county_fips"], c["state_fips"], c["county_name"],
             "no_county_government", 0, _now_iso()),
        )
        n_counties += 1

    munis = fetch_census_municipalities(from_file=munis_from_file)
    n_munis = 0
    for m in munis:
        conn.execute(
            """
            INSERT OR IGNORE INTO municipalities
                (state_fips, county_fips, muni_name, muni_type,
                 population, official_website, last_verified_at)
            VALUES (?, ?, ?, ?, ?, NULL, ?)
            """,
            (m["state_fips"], m["county_fips"],
             m["muni_name"], m["muni_type"], m["population"], _now_iso()),
        )
        n_munis += 1

    conn.commit()
    return {"counties": n_counties, "municipalities": n_munis}


# ---------------------------------------------------------------------------
# Officials data fetch (data lake + collection_log)
# ---------------------------------------------------------------------------

def _utc_stamp() -> str:
    """ISO-8601-style UTC timestamp safe for filenames (no colons)."""
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%SZ")


def _now_iso() -> str:
    """Current UTC time as a full ISO-8601 string for collected_at / last_verified_at."""
    return datetime.now(timezone.utc).isoformat()


def _resolve_source_id(conn: sqlite3.Connection, source_name: str) -> int:
    """Look up a sources.source_id by canonical name. Errors if seed missing."""
    row = conn.execute(
        "SELECT source_id FROM sources WHERE source_name = ?", (source_name,)
    ).fetchone()
    if not row:
        raise RuntimeError(f"source not seeded: {source_name!r}")
    return row[0]


def _last_successful_hash(conn: sqlite3.Connection, source_id: int) -> str | None:
    """Most recent raw_content_hash for a successful fetch from this source."""
    row = conn.execute(
        "SELECT raw_content_hash FROM collection_log "
        "WHERE source_id = ? AND status = 'success' "
        "  AND raw_content_hash IS NOT NULL "
        "ORDER BY run_at DESC LIMIT 1",
        (source_id,),
    ).fetchone()
    return row[0] if row else None


def _last_storage_path(conn: sqlite3.Connection, source_id: int) -> str | None:
    """Most recent storage_path for this source - returned when the hash
    matches, so the caller can re-parse the existing artifact instead of
    re-writing identical bytes to the lake."""
    row = conn.execute(
        "SELECT storage_path FROM collection_log "
        "WHERE source_id = ? AND status = 'success' "
        "  AND storage_path IS NOT NULL "
        "ORDER BY run_at DESC LIMIT 1",
        (source_id,),
    ).fetchone()
    return row[0] if row else None


def fetch_to_data_lake(
    url: str, source_slug: str, conn: sqlite3.Connection, source_id: int
) -> tuple[Path, int]:
    """HTTP GET, write append-only to data/raw/<slug>/<UTC>.<ext>, log.

    Hash-based change detection: if the SHA-256 of the response body
    matches the most recent successful fetch for this source, the
    pipeline logs status='skipped_unchanged' and returns the existing
    storage_path without writing a duplicate copy. The Section-3
    "monthly hash-check" optimization, scaled down to the MVP.

    Returns (path, log_id) - the log_id is used by `update_log_counts`
    to back-fill records_found/records_updated after reconcile.
    """
    try:
        resp = requests.get(
            url, headers={"User-Agent": USER_AGENT}, timeout=HTTP_TIMEOUT
        )
        resp.raise_for_status()
    except Exception as exc:  # noqa: BLE001
        conn.execute(
            "INSERT INTO collection_log "
            "(source_id, run_at, status, error_message) VALUES (?, ?, 'failed', ?)",
            (source_id, _now_iso(), f"{type(exc).__name__}: {exc}"),
        )
        conn.commit()
        raise

    raw = resp.content
    sha = hashlib.sha256(raw).hexdigest()
    last_hash = _last_successful_hash(conn, source_id)

    # Hash-skip path: same bytes as last successful run.
    if last_hash and last_hash == sha:
        existing = _last_storage_path(conn, source_id)
        cur = conn.execute(
            "INSERT INTO collection_log "
            "(source_id, run_at, status, storage_path, raw_content_hash) "
            "VALUES (?, ?, 'skipped_unchanged', ?, ?)",
            (source_id, _now_iso(), existing, sha),
        )
        log_id = cur.lastrowid
        conn.execute(
            "UPDATE sources SET last_fetched_at = ? WHERE source_id = ?",
            (_now_iso(), source_id),
        )
        conn.commit()
        return (Path(existing) if existing else Path()), log_id

    ct = resp.headers.get("Content-Type", "").lower()
    ext = "json" if "json" in ct else (
        "html" if "html" in ct else (Path(url).suffix.lstrip(".") or "bin")
    )
    path = RAW_DIR / source_slug / f"{_utc_stamp()}.{ext}"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(raw)

    cur = conn.execute(
        "INSERT INTO collection_log "
        "(source_id, run_at, status, storage_path, raw_content_hash) "
        "VALUES (?, ?, 'success', ?, ?)",
        (source_id, _now_iso(), str(path), sha),
    )
    log_id = cur.lastrowid
    conn.execute(
        "UPDATE sources SET last_fetched_at = ? WHERE source_id = ?",
        (_now_iso(), source_id),
    )
    conn.commit()
    return path, log_id


def ingest_local_file(
    src_path: str | Path, source_slug: str,
    conn: sqlite3.Connection, source_id: int,
) -> tuple[Path, int]:
    """Treat a manually-downloaded file as a fetch.

    Same provenance rules as fetch_to_data_lake. The file is COPIED into
    the lake so storage_path always points there. Returns (path, log_id)
    so the caller can later UPDATE the same collection_log row with
    records_found / records_updated once parsing + reconciliation finish.
    """
    src = Path(src_path)
    raw = src.read_bytes()
    sha = hashlib.sha256(raw).hexdigest()
    ext = src.suffix.lstrip(".") or "bin"
    path = RAW_DIR / source_slug / f"{_utc_stamp()}.{ext}"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(raw)
    cur = conn.execute(
        "INSERT INTO collection_log "
        "(source_id, run_at, status, storage_path, raw_content_hash, error_message) "
        "VALUES (?, ?, 'success', ?, ?, 'ingested from local file')",
        (source_id, _now_iso(), str(path), sha),
    )
    log_id = cur.lastrowid
    conn.execute(
        "UPDATE sources SET last_fetched_at = ? WHERE source_id = ?",
        (_now_iso(), source_id),
    )
    conn.commit()
    return path, log_id


def _to_url_slug(muni_name: str) -> str:
    """Convert a muni display name to its URL filename stem.

    RI SOS publishes per-muni JSON at <base>/<slug>.json with the
    convention: lowercase, spaces -> underscores. Verified against
    single-word (barrington.json) and multi-word (north_smithfield.json)
    URLs. Per-muni 404s still log status='failed' so the loop keeps
    going if SOS ever changes the convention.
    """
    return muni_name.lower().replace(" ", "_")


def fetch_ri_sos_all_munis(
    conn: sqlite3.Connection, source_id: int, base_url: str,
) -> list[tuple[str, Path, int]]:
    """Loop over every seeded RI municipality and fetch its results JSON
    into the data lake.

    Per-muni URL = base_url/<slug>.json. Each fetch goes through
    `fetch_to_data_lake`, so each muni gets its own collection_log row
    with hash + storage_path - no special-cased provenance for the loop.

    Returns [(muni_name, lake_path, log_id), ...] for successful fetches.
    Failed fetches log status='failed' and are skipped; the loop keeps
    going so one bad slug or one missing per-muni file doesn't sink the
    whole run.
    """
    munis = [r[0] for r in conn.execute(
        "SELECT muni_name FROM municipalities "
        "WHERE state_fips = ? ORDER BY muni_name",
        (STATE_FIPS,),
    )]
    out: list[tuple[str, Path, int]] = []
    base = base_url.rstrip("/")
    for muni in munis:
        slug = _to_url_slug(muni)
        url = f"{base}/{slug}.json"
        try:
            path, log_id = fetch_to_data_lake(url, TIER1_SLUG, conn, source_id)
            out.append((muni, path, log_id))
        except Exception as exc:  # noqa: BLE001
            print(f"  fetch failed for {muni!r} ({url}): {exc}",
                  file=sys.stderr)
    return out


def update_log_counts(
    conn: sqlite3.Connection, log_id: int,
    records_found: int, records_updated: int,
) -> None:
    """Back-fill records_found / records_updated on a collection_log row
    once the parse + reconcile stages finish. Lets a future query of
    `collection_log` answer 'how many records did this run produce?'
    without re-parsing the artifact."""
    conn.execute(
        "UPDATE collection_log SET records_found = ?, records_updated = ? "
        "WHERE log_id = ?",
        (records_found, records_updated, log_id),
    )
    conn.commit()


# ---------------------------------------------------------------------------
# Parser (Tier 1)
# ---------------------------------------------------------------------------

# RI SOS JSON shape (one file per municipality):
#   { "election_date": "...",
#     "contests": [
#       { "name": "Town Clerk TOWN OF BRISTOL",
#         "votes_allowed": "1",                    # number of seats up
#         "candidates": [
#           {"name": "DEM Melissa M. Cordeiro", "party_code": "DEM", "votes": "9086"},
#           {"name": "Write-in",                "party_code": "NON", "votes": "273"} ]},
#       ...]}
#
# The parser:
#   - filters contests to those whose name contains "TOWN OF" or "CITY OF"
#     (drops Presidential, Senator in Congress, General Assembly seats,
#     ballot questions like "1. CONSTITUTIONAL CONVENTION", etc.)
#   - splits the contest name into (office_title, muni_name)
#   - sorts candidates by votes descending, drops "Write-in" entries,
#     takes the top `votes_allowed` as winners (the design's max-votes
#     winner-detection logic for sources that don't tag winners explicitly)
#   - strips the party-code prefix that RI SOS encodes in the displayed
#     candidate name: "DEM Melissa M. Cordeiro" -> "Melissa M. Cordeiro"
#   - canonicalizes party_code (DEM -> Democratic, etc.)

# Match a "TOWN OF X" or "CITY OF X" suffix on a contest name.
_MUNICIPAL_CONTEST_RE = re.compile(
    r"\b(?:TOWN|CITY)\s+OF\s+(?P<muni>.+?)\s*$",
    re.IGNORECASE,
)

# party_code -> canonical party label. Unknown codes pass through verbatim
# so an analyst can spot them and extend the map.
_PARTY_CANONICAL = {
    "DEM": "Democratic", "REP": "Republican",
    "Ind": "Independent", "IND": "Independent",
    "NON": "Nonpartisan",
    "Lib": "Libertarian", "LIB": "Libertarian",
    "Grn": "Green",       "GRE": "Green",
    "Mod": "Moderate",    "MOD": "Moderate",
    "S&L": "Socialism and Liberation",
    "Ken": "Independent (Kennedy)",
    "Par": "Party for the Restoration",
}


def _split_municipal_contest(contest_name: str) -> tuple[str | None, str | None]:
    """Return (office_title, muni_name) if `contest_name` is municipal,
    else (None, None). Examples:
        "Town Clerk TOWN OF BRISTOL"             -> ("Town Clerk", "Bristol")
        "Mayor CITY OF CRANSTON"                 -> ("Mayor", "Cranston")
        "Senator in General Assembly District 10" -> (None, None)
        "1. CONSTITUTIONAL CONVENTION"           -> (None, None)
    """
    if not contest_name:
        return None, None
    m = _MUNICIPAL_CONTEST_RE.search(contest_name)
    if not m:
        return None, None
    muni_raw = m.group("muni").strip()
    muni_name = muni_raw.title()  # "BRISTOL" -> "Bristol", "EAST GREENWICH" -> "East Greenwich"
    office_title = contest_name[:m.start()].strip()
    return office_title, muni_name


def _strip_party_prefix(name: str | None, party_code: str | None) -> str:
    """Remove a leading "<party_code> " from a candidate's displayed name."""
    if not name:
        return ""
    if not party_code:
        return name.strip()
    prefix = f"{party_code} "
    return name[len(prefix):].strip() if name.startswith(prefix) else name.strip()


def _canonical_party(party_code: str | None) -> str | None:
    """Map a raw party_code (DEM, REP, Ind, NON, Lib...) to a canonical
    label. Unknown codes pass through verbatim so analysts can extend the map."""
    if not party_code:
        return None
    return _PARTY_CANONICAL.get(party_code.strip(), party_code.strip())


def _is_write_in(name: str | None) -> bool:
    """True if a candidate's display name marks them as a write-in row."""
    return bool(name) and "write-in" in name.lower()


def parse_ri_sos_winners(
    path: str | Path,
    *,
    source_url: str | None = None,
) -> list[dict]:
    """Parse one RI SOS per-municipality results JSON.

    Returns a list of winner records (one per (contest, seat) winning
    candidate) that match the validation contract. The validator decides
    which records to load vs flag - this parser is intentionally
    permissive about its output, since silently dropping records is
    how parsers go bad without anyone noticing.

    Every record carries `storage_path` so a downstream `needs_review`
    row can trace back to the lake artifact.
    """
    src_path = Path(path)
    data = json.loads(src_path.read_text(encoding="utf-8"))
    storage_path = str(src_path)
    out_url = source_url or TIER1_URL
    # The SOS payload carries the election date at the top level
    # (e.g. "November 05, 2024"). Stamp it on every record so downstream
    # term_start derivation has the data and the CLI doesn't need a flag.
    election_date = data.get("election_date")

    out: list[dict] = []
    for contest in data.get("contests", []):
        contest_name = contest.get("name", "")
        office_title, muni_name = _split_municipal_contest(contest_name)
        if not office_title:
            continue   # state, federal, or ballot-question contest

        # Number of seats up. Defaults to 1 for malformed input.
        try:
            seats = int(str(contest.get("votes_allowed", "1")).strip() or "1")
        except ValueError:
            seats = 1

        # Rank candidates by votes (desc), excluding write-ins.
        ranked: list[tuple[int, dict]] = []
        for cand in contest.get("candidates", []):
            if _is_write_in(cand.get("name")):
                continue
            try:
                votes = int(str(cand.get("votes", "0")).strip() or "0")
            except ValueError:
                votes = 0
            ranked.append((votes, cand))
        ranked.sort(key=lambda t: -t[0])

        # Top-N winners.
        for _votes, cand in ranked[:seats]:
            party_code = cand.get("party_code")
            full_name = _strip_party_prefix(cand.get("name"), party_code)
            out.append({
                "muni_name":     muni_name,
                "full_name":     full_name,
                "office_title":  office_title,
                "party":         _canonical_party(party_code),
                "source_url":    out_url,
                "storage_path":  storage_path,
                "election_date": election_date,
            })
    return out


# ---------------------------------------------------------------------------
# Validator + needs_review
# ---------------------------------------------------------------------------

REQUIRED_FIELDS = ("muni_name", "full_name", "office_title")


def validate_record(
    record: dict, conn: sqlite3.Connection
) -> tuple[bool, str | None, int | None]:
    """Per-record contract: required fields + muni resolution.

    Returns (is_valid, reason_if_invalid, muni_id_if_valid).
    Case-insensitive muni match handles minor source-side spelling drift.
    """
    for f in REQUIRED_FIELDS:
        if not record.get(f) or not str(record[f]).strip():
            return False, f"missing or empty: {f}", None

    row = conn.execute(
        "SELECT muni_id FROM municipalities "
        "WHERE state_fips = ? AND LOWER(muni_name) = LOWER(?)",
        (STATE_FIPS, record["muni_name"].strip()),
    ).fetchone()
    if not row:
        return False, f"unknown municipality: {record['muni_name']!r}", None
    return True, None, row[0]


def write_to_needs_review(
    conn: sqlite3.Connection, record: dict, reason: str, source_id: int | None,
) -> None:
    """Persist a failing record. storage_path on the record (set by the
    parser) becomes the audit trail back to the lake artifact."""
    conn.execute(
        "INSERT INTO needs_review "
        "(raw_record, failure_reason, source_id, storage_path, flagged_at) "
        "VALUES (?, ?, ?, ?, ?)",
        (json.dumps(record, default=str), reason, source_id,
         record.get("storage_path"), _now_iso()),
    )
    conn.commit()


# ---------------------------------------------------------------------------
# Reconciler (fuzzy)
# ---------------------------------------------------------------------------

def fuzzy_canonical_office_type(
    local_title: str, conn: sqlite3.Connection,
    threshold: int = FUZZY_TITLE_THRESHOLD,
) -> tuple[int | None, str | None, float]:
    """Find canonical office_type by fuzzy-matching the raw title.

    Joins office_types -> office_type_synonyms, computes
    rapidfuzz.token_set_ratio against every (canonical, synonym) pair,
    returns the highest scorer above threshold.

    token_set_ratio is the right scorer here: word-set similarity
    handles the common drift cases ("Town Council" vs "Town Council
    Member") without over-rewarding substring matches the way
    partial_ratio would.
    """
    if not local_title:
        return None, None, 0.0

    rows = conn.execute(
        """
        SELECT ot.office_type_id, ot.canonical_name, syn.synonym
        FROM office_types ot
        JOIN office_type_synonyms syn USING(office_type_id)
        """
    ).fetchall()
    if not rows:
        return None, None, 0.0

    best_id, best_name, best_score = None, None, 0.0
    title_lc = local_title.lower()
    for ot_id, canonical, synonym in rows:
        score = fuzz.token_set_ratio(title_lc, synonym.lower())
        if score > best_score:
            best_id, best_name, best_score = ot_id, canonical, score
    if best_score >= threshold:
        return best_id, best_name, best_score
    return None, None, best_score


def fuzzy_match_official_within_muni(
    full_name: str, muni_id: int, conn: sqlite3.Connection,
    threshold: int = FUZZY_NAME_THRESHOLD,
) -> int | None:
    """Find an existing official in this muni by fuzzy name similarity.

    Scoping to the muni keeps a "John Smith" in Providence from
    colliding with a "John Smith" in Pawtucket. Above-threshold matches
    are reused; below-threshold returns None and the caller creates a
    new officials row. Conservative threshold favors splits over
    false merges.
    """
    rows = conn.execute(
        """
        SELECT DISTINCT off.official_id, off.full_name
        FROM officials off
        JOIN office_holders oh ON oh.official_id = off.official_id
        JOIN offices       o  ON o.office_id     = oh.office_id
        WHERE o.muni_id = ?
        """,
        (muni_id,),
    ).fetchall()
    if not rows:
        return None
    names = [r[1] for r in rows]
    match = process.extractOne(full_name, names, scorer=fuzz.ratio)
    if match and match[1] >= threshold:
        return rows[names.index(match[0])][0]
    return None


def _find_or_create_office(
    conn: sqlite3.Connection, *, muni_id: int, office_type_id: int,
    local_title: str, district_or_seat: str | None,
) -> int:
    """Look up an offices row by its UNIQUE key, inserting if missing.

    Handles SQLite's NULL-vs-NULL inequality quirk on the district_or_seat
    column (NULLs aren't equal under UNIQUE), so the find branch picks
    IS-NULL when no seat is provided.
    """
    if district_or_seat is None:
        existing = conn.execute(
            "SELECT office_id FROM offices "
            "WHERE muni_id = ? AND office_type_id = ? "
            "AND local_title = ? AND district_or_seat IS NULL",
            (muni_id, office_type_id, local_title),
        ).fetchone()
    else:
        existing = conn.execute(
            "SELECT office_id FROM offices "
            "WHERE muni_id = ? AND office_type_id = ? "
            "AND local_title = ? AND district_or_seat = ?",
            (muni_id, office_type_id, local_title, district_or_seat),
        ).fetchone()
    if existing:
        return existing[0]
    cur = conn.execute(
        "INSERT INTO offices "
        "(muni_id, office_type_id, local_title, district_or_seat) "
        "VALUES (?, ?, ?, ?)",
        (muni_id, office_type_id, local_title, district_or_seat),
    )
    return cur.lastrowid


def _split_name(full_name: str) -> tuple[str | None, str | None]:
    """Cheap first/last split. Last whitespace token is last_name.

    Real entity resolution would handle suffixes, hyphens, surname-first
    cultures, etc. - that's the v2 placeholder noted in the design doc's
    Section 5 entity-resolution layer.
    """
    parts = full_name.strip().split()
    if not parts:
        return None, None
    if len(parts) == 1:
        return None, parts[0]
    return " ".join(parts[:-1]), parts[-1]


def _today_iso() -> str:
    """Current UTC date as YYYY-MM-DD - used as the term_end value when
    the reconciler closes a previous holder on a person change."""
    return datetime.now(timezone.utc).date().isoformat()


def reconcile_record(
    record: dict, conn: sqlite3.Connection, *,
    source_id: int, term_start: str,
) -> str:
    """The four moves from Section 3 of the design doc.

    Returns one of:
      'needs_review'      - validation failed or no canonical match
      'loaded'            - new office_holders row inserted (no prior current)
      'closed_and_opened' - prior current single-seat row closed,
                            new current row opened (person changed)
      'touched'           - existing current row's last_verified_at refreshed
    """
    now = _now_iso()

    # Move 1: per-record validation.
    ok, reason, muni_id = validate_record(record, conn)
    if not ok:
        write_to_needs_review(conn, record, reason or "validation failed", source_id)
        return "needs_review"

    # Move 2: canonical title mapping (fuzzy).
    ot_id, canonical, score = fuzzy_canonical_office_type(record["office_title"], conn)
    if ot_id is None:
        write_to_needs_review(
            conn, record,
            f"no canonical office_type for {record['office_title']!r} "
            f"(best fuzzy score {score:.0f} < threshold {FUZZY_TITLE_THRESHOLD})",
            source_id,
        )
        return "needs_review"

    is_typically_board = bool(conn.execute(
        "SELECT is_typically_board FROM office_types WHERE office_type_id = ?",
        (ot_id,),
    ).fetchone()[0])

    office_id = _find_or_create_office(
        conn,
        muni_id=muni_id, office_type_id=ot_id,
        local_title=record["office_title"],
        district_or_seat=record.get("district_or_seat"),
    )

    # Move 3: person matching (fuzzy, scoped per-muni).
    official_id = fuzzy_match_official_within_muni(record["full_name"], muni_id, conn)
    if official_id is None:
        first, last = _split_name(record["full_name"])
        cur = conn.execute(
            "INSERT INTO officials (full_name, first_name, last_name, party) "
            "VALUES (?, ?, ?, ?)",
            (record["full_name"], first, last, record.get("party")),
        )
        official_id = cur.lastrowid

    # Move 4: term lifecycle. Behavior differs by single-seat vs board:
    #   single-seat: at most one is_current=1 per office. If a different
    #                person currently holds it, close that row (set
    #                is_current=0, term_end=today) and open a new one.
    #                If the same person, touch.
    #   board:       keyed by (office_id, official_id). One is_current
    #                row per person per seat. Same-person re-runs touch;
    #                new people open a new row alongside any existing ones.
    # This is the pseudocode from Section 3 of the design doc.
    if is_typically_board:
        existing = conn.execute(
            "SELECT holder_id FROM office_holders "
            "WHERE office_id = ? AND official_id = ? AND is_current = 1",
            (office_id, official_id),
        ).fetchone()
        if existing:
            conn.execute(
                "UPDATE office_holders SET last_verified_at = ?, source_id = ? "
                "WHERE holder_id = ?",
                (now, source_id, existing[0]),
            )
            conn.commit()
            return "touched"
    else:
        existing = conn.execute(
            "SELECT holder_id, official_id FROM office_holders "
            "WHERE office_id = ? AND is_current = 1",
            (office_id,),
        ).fetchone()
        if existing and existing[1] == official_id:
            conn.execute(
                "UPDATE office_holders SET last_verified_at = ?, source_id = ? "
                "WHERE holder_id = ?",
                (now, source_id, existing[0]),
            )
            conn.commit()
            return "touched"
        if existing:
            # Different person in the seat - close the old row.
            conn.execute(
                "UPDATE office_holders "
                "SET is_current = 0, term_end = ? WHERE holder_id = ?",
                (_today_iso(), existing[0]),
            )
            conn.execute(
                "INSERT INTO office_holders "
                "(official_id, office_id, term_start, is_current, "
                " assumption_method, source_id, collected_at, last_verified_at) "
                "VALUES (?, ?, ?, 1, 'elected', ?, ?, ?)",
                (official_id, office_id, term_start, source_id, now, now),
            )
            conn.commit()
            return "closed_and_opened"

    # No existing current row matched - open a new one.
    conn.execute(
        "INSERT INTO office_holders "
        "(official_id, office_id, term_start, is_current, "
        " assumption_method, source_id, collected_at, last_verified_at) "
        "VALUES (?, ?, ?, 1, 'elected', ?, ?, ?)",
        (official_id, office_id, term_start, source_id, now, now),
    )
    conn.commit()
    return "loaded"


def reconcile_records(
    records: list[dict], conn: sqlite3.Connection, *,
    source_id: int, term_start: str,
) -> dict[str, int]:
    """Apply reconcile_record to a batch and return aggregate outcome counts."""
    counts = {"loaded": 0, "closed_and_opened": 0, "touched": 0, "needs_review": 0}
    for record in records:
        result = reconcile_record(
            record, conn, source_id=source_id, term_start=term_start
        )
        counts[result] = counts.get(result, 0) + 1
    return counts


# ---------------------------------------------------------------------------
# Post-load sanity checks
# ---------------------------------------------------------------------------
# Section 4 of the design doc names two checks for v1:
#   - gap_rate: counties (or, here, munis) with elected gov but zero
#     officeholders. Catches a parser silently returning an empty list
#     or a reconciler dropping rows.
#   - staleness: is_current rows with last_verified_at older than a
#     threshold. Catches a source going dark.
# Section 4 lists a third category ("cross-entity consistency": no
# duplicate-current on single-seat, term_start non-NULL, etc.) explicitly
# under "would add with more time" in the design doc itself, so those
# stay deferred and are noted as such in the README.

# Thresholds. Tuneable per state in production.
GAP_RATE_THRESHOLD  = 0.30   # >30% of expected munis empty -> pipeline broken
STALENESS_DAYS      = 90     # current rows older than this are stale
STALENESS_THRESHOLD = 0.20   # >20% stale -> a source is going dark


def check_gap_rate_munis(conn: sqlite3.Connection) -> dict:
    """Fraction of RI municipalities with zero current officeholders."""
    total = conn.execute(
        "SELECT COUNT(*) FROM municipalities WHERE state_fips = ?",
        (STATE_FIPS,),
    ).fetchone()[0]
    if total == 0:
        return {"name": "gap_rate_munis", "passed": True,
                "detail": "no munis seeded; nothing to check", "metric": 0.0}
    munis_with_holders = conn.execute(
        """
        SELECT COUNT(DISTINCT m.muni_id)
        FROM municipalities m
        JOIN offices         o  ON o.muni_id = m.muni_id
        JOIN office_holders  oh ON oh.office_id = o.office_id AND oh.is_current = 1
        WHERE m.state_fips = ?
        """,
        (STATE_FIPS,),
    ).fetchone()[0]
    gap = (total - munis_with_holders) / total
    passed = gap <= GAP_RATE_THRESHOLD
    return {
        "name": "gap_rate_munis",
        "passed": passed,
        "metric": gap,
        "detail": (
            f"{total - munis_with_holders}/{total} munis have zero current "
            f"officeholders (gap {gap:.1%}, threshold {GAP_RATE_THRESHOLD:.0%})"
        ),
    }


def check_staleness(conn: sqlite3.Connection) -> dict:
    """Fraction of is_current rows whose last_verified_at is too old."""
    total = conn.execute(
        "SELECT COUNT(*) FROM office_holders WHERE is_current = 1"
    ).fetchone()[0]
    if total == 0:
        return {"name": "staleness", "passed": True,
                "detail": "no current holders; nothing to check", "metric": 0.0}
    cutoff = (datetime.now(timezone.utc)
              - timedelta(days=STALENESS_DAYS)).isoformat()
    stale = conn.execute(
        """
        SELECT COUNT(*) FROM office_holders
        WHERE is_current = 1
          AND (last_verified_at IS NULL OR last_verified_at < ?)
        """,
        (cutoff,),
    ).fetchone()[0]
    rate = stale / total
    return {
        "name": "staleness",
        "passed": rate <= STALENESS_THRESHOLD,
        "metric": rate,
        "detail": (
            f"{stale}/{total} current holders last verified more than "
            f"{STALENESS_DAYS} days ago "
            f"(rate {rate:.1%}, threshold {STALENESS_THRESHOLD:.0%})"
        ),
    }


def run_post_load_checks(conn: sqlite3.Connection) -> list[dict]:
    """Run every post-load sanity check; return a list of result dicts."""
    return [check_gap_rate_munis(conn), check_staleness(conn)]


# ---------------------------------------------------------------------------
# Tier 2 cross-check
# ---------------------------------------------------------------------------

# RI Open Data Portal labels precincts as "<Muni> <precinct token>", where
# the precinct token is a 4-digit number (e.g. "Barrington 0101"), the
# string "Limited" (limited-ballot precincts), or "Presidential" (federal-
# only precincts). The "Federal Precinct #X" rows are not municipalities
# and are skipped.
_PRECINCT_SUFFIX_TOKENS = {"limited", "presidential"}


def _extract_muni_from_precinct_label(label: str | None) -> str | None:
    """Strip trailing precinct tokens from a "<Muni> <precinct>" label.

    Iterative: handles single tokens ("Barrington 0101" -> "Barrington")
    and multi tokens ("Providence Limited 2" -> "Providence"). Returns
    None for non-municipal rows (e.g. "Federal Precinct #1").
    """
    if not label:
        return None
    if str(label).startswith("Federal Precinct"):
        return None
    out = str(label)
    while True:
        parts = out.rsplit(" ", 1)
        if len(parts) != 2:
            break
        if parts[1].isdigit() or parts[1].lower() in _PRECINCT_SUFFIX_TOKENS:
            out = parts[0]
        else:
            break
    return out.strip() or None


def _muni_set_from_xlsx(content_or_path) -> set[str]:
    """Extract the unique muni names from an RI Open Data Portal XLSX.

    The portal's election summaries key every row by a "City/Town - Precinct"
    column on the first sheet. We walk distinct values in that column,
    strip the precinct suffix, and lowercase for case-insensitive matching
    against the loaded munis.
    """
    from io import BytesIO
    if isinstance(content_or_path, (bytes, bytearray)):
        wb = load_workbook(BytesIO(bytes(content_or_path)),
                           read_only=True, data_only=True)
    else:
        wb = load_workbook(filename=str(content_or_path),
                           read_only=True, data_only=True)
    out: set[str] = set()
    # First sheet is conventionally Reg_Voters in the RI portal export;
    # any sheet keyed by the same first column works equally well.
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        muni = _extract_muni_from_precinct_label(row[0])
        if muni:
            out.add(muni.lower())
    return out


def _muni_set_from_json(raw: object) -> set[str]:
    """Extract muni names from a JSON Tier 2 list (legacy/test path)."""
    if isinstance(raw, dict):
        for k in ("results", "data", "records", "items"):
            if isinstance(raw.get(k), list):
                raw = raw[k]
                break
    if not isinstance(raw, list):
        raise ValueError(
            f"Tier 2 JSON is not a list: type={type(raw).__name__}"
        )
    out: set[str] = set()
    for entry in raw:
        if not isinstance(entry, dict):
            continue
        name = (
            entry.get("name") or entry.get("muni_name")
            or entry.get("municipality") or entry.get("city")
            or entry.get("town")
        )
        if name:
            out.add(str(name).strip().lower())
    return out


def cross_check_munis(
    conn: sqlite3.Connection, tier2_path_or_url: str
) -> dict[str, int]:
    """Validate loaded munis against a Tier 2 reference list.

    Accepts either:
      - a JSON file/URL: list of {name|muni_name|municipality|city|town: ...}
      - an XLSX file/URL: RI Open Data Portal election summary, with a
        "City/Town - Precinct" column on the first sheet
        (e.g. Gen24EX.xlsx, ~480 precinct rows aggregating to 39 munis)

    Set-membership semantics: every muni that owns at least one current
    `office_holder` in our Tier 1 load must appear in the Tier 2
    reference. Mismatches are printed to stderr; counts are returned.

    Richer cross-checks the XLSX makes possible (per-muni vote totals,
    per-(muni, contest) winner verification using `Candidate_Breakout`,
    contest-coverage parity) are tracked in the with-more-time list.
    """
    src = str(tier2_path_or_url)
    is_url = src.startswith(("http://", "https://"))
    is_xlsx = src.lower().endswith((".xlsx", ".xlsm"))

    if is_url:
        # We need the bytes for XLSX (random-access format) and JSON alike.
        resp = requests.get(
            src, headers={"User-Agent": USER_AGENT}, timeout=HTTP_TIMEOUT
        )
        resp.raise_for_status()
        ct = resp.headers.get("Content-Type", "").lower()
        # Treat as XLSX if the URL says so OR the Content-Type points at
        # an Office Open XML body. Otherwise assume JSON.
        is_xlsx = is_xlsx or "spreadsheet" in ct or "officedocument" in ct
        if is_xlsx:
            canonical_names = _muni_set_from_xlsx(resp.content)
        else:
            canonical_names = _muni_set_from_json(resp.json())
    else:
        if is_xlsx:
            canonical_names = _muni_set_from_xlsx(src)
        else:
            canonical_names = _muni_set_from_json(
                json.loads(Path(src).read_text(encoding="utf-8"))
            )

    loaded = [r[0] for r in conn.execute(
        """
        SELECT DISTINCT m.muni_name FROM municipalities m
        JOIN offices       o  ON o.muni_id     = m.muni_id
        JOIN office_holders oh ON oh.office_id = o.office_id AND oh.is_current = 1
        WHERE m.state_fips = ?
        ORDER BY m.muni_name
        """,
        (STATE_FIPS,),
    )]

    matched = unmatched = 0
    for name in loaded:
        if name.lower() in canonical_names:
            matched += 1
        else:
            print(f"  Tier 2 mismatch: {name!r} not in reference", file=sys.stderr)
            unmatched += 1
    print(
        f"Cross-check vs Tier 2 ({TIER2_NAME}): "
        f"{matched} matched / {unmatched} unmatched / "
        f"{len(canonical_names)} reference entries"
    )
    return {"matched": matched, "unmatched": unmatched,
            "reference_size": len(canonical_names)}


# ---------------------------------------------------------------------------
# Unit test (one example)
# ---------------------------------------------------------------------------
# A single pytest-discoverable test as proof-of-pattern. Run with:
#
#     pytest ri_officials_mvp.py or python -m pytest ri_officials_mvp.py
#
# Mirrors the parse_nc_sboe_returns_officials test from Section 4 of the
# Part 1 design doc - same shape, same contract assertions. The MVP keeps
# one example here; the with-more-time section of the README enumerates
# the additional unit-test cases this would expand to (validator dirty
# inputs, fuzzy match boundaries, _split_cousub_name regex fallback,
# term lifecycle across cycles, date-rollover logic in
# _term_start_from_election_date). The full 26-test suite for the same
# pipeline lives at reference_implementation/tests/.

def test_parse_ri_sos_winners_filters_losers_and_extracts_contract_fields():
    """Parser contract test parallel to parse_nc_sboe_returns_officials.

    INLINE_TIER1 has 7 contests in real RI SOS shape:
      - 1 state-level contest (Senator in General Assembly) -> filtered
      - 1 ballot question (CONSTITUTIONAL CONVENTION) -> filtered
      - 1 single-seat municipal (Mayor PVD, votes_allowed=1) -> 1 winner
      - 1 multi-seat municipal (City Council PVD, votes_allowed=3) -> 3 winners
      - 1 single-seat municipal (Town Clerk MID, votes_allowed=1) -> 1 winner
      - 1 multi-seat municipal (School Committee NPT, votes_allowed=2) -> 2 winners
      - 1 unknown-muni municipal (Mayor ATL, validator will reject) -> 1 winner

    Parser MUST: filter non-municipal contests, drop write-ins, take top
    `votes_allowed` candidates per contest, strip the party-code prefix
    from each candidate name ("DEM Brett Smiley" -> "Brett Smiley"),
    canonicalize the party_code (DEM -> Democratic), and attach
    storage_path for the audit trail.
    """
    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "tier1.json"
        path.write_text(INLINE_TIER1)
        records = parse_ri_sos_winners(path)

    # 1 + 3 + 1 + 2 + 1 = 8 winner records emitted.
    assert len(records) == 8
    assert all(r["muni_name"] for r in records)
    assert all(r["office_title"] for r in records)
    assert all(r["full_name"] for r in records)
    assert all(r["storage_path"] for r in records)

    # Sub-top-votes-allowed candidates dropped.
    assert not any(r["full_name"] == "Some Loser" for r in records)
    assert not any(r["full_name"] == "Fourth Loser" for r in records)
    assert not any(r["full_name"] == "Sad Loser" for r in records)
    # Write-ins dropped.
    assert not any("write-in" in r["full_name"].lower() for r in records)
    # Party prefix stripped.
    assert not any(r["full_name"].startswith(("DEM ", "REP ", "Ind "))
                   for r in records)
    # Specific winner round-trips with the prefix removed.
    assert any(r["full_name"] == "Brett Smiley" and r["party"] == "Democratic"
               for r in records)
    # State / federal contest dropped.
    assert not any("Senator" in r["office_title"] for r in records)
    # Multi-seat: 3 City Council winners.
    council = [r for r in records if r["office_title"] == "City Council"]
    assert len(council) == 3


# ---------------------------------------------------------------------------
# Self-test (offline, deterministic)
# ---------------------------------------------------------------------------
# Inline Census-format fixtures: 2 counties, 4 cousubs - enough to
# exercise the fetch loop and the cross-state title fuzzy match without
# requiring network. The munis fixture is intentionally a strict subset
# of RI's real cousubs so the cross-check still produces clean matches.

INLINE_CENSUS_COUNTIES = """[
  ["NAME","state","county"],
  ["Providence County, Rhode Island","44","007"],
  ["Newport County, Rhode Island","44","005"]
]"""

INLINE_CENSUS_COUSUBS = """[
  ["NAME","P1_001N","state","county","county subdivision"],
  ["Providence city, Providence County, Rhode Island","190934","44","007","59000"],
  ["Pawtucket city, Providence County, Rhode Island","75604","44","007","54640"],
  ["Newport city, Newport County, Rhode Island","25163","44","005","49960"],
  ["Middletown town, Newport County, Rhode Island","17075","44","005","45040"]
]"""

# Inline mini RI SOS payload that mirrors the real per-muni JSON shape:
# `contests` array, `votes_allowed` for seat count, party encoded both
# as a prefix on the candidate's displayed name and as `party_code`,
# write-ins as candidate rows the parser must drop. Includes the dirty
# cases the parser/validator/reconciler should catch:
#   - a state contest (Senator in General Assembly) -> parser filters
#   - a ballot question (CONSTITUTIONAL CONVENTION) -> parser filters
#   - a contest pointing at an unseeded muni (Atlantis) -> validator routes to needs_review
#   - a multi-seat race (City Council) -> reconciler picks top 3 by votes
#   - a single-seat race with a "Some Loser" -> parser keeps only top 1
INLINE_TIER1 = """{
  "election_name": "General Election",
  "election_date": "November 05, 2024",
  "contests": [
    { "name": "Senator in General Assembly District 10",
      "votes_allowed": "1",
      "candidates": [
        {"name": "DEM Walter S. Felag, Jr.", "party_code": "DEM", "votes": "2203"},
        {"name": "REP Allyn E. Meyers",      "party_code": "REP", "votes": "1710"}
      ]
    },
    { "name": "Mayor CITY OF PROVIDENCE",
      "votes_allowed": "1",
      "candidates": [
        {"name": "DEM Brett Smiley", "party_code": "DEM", "votes": "26500"},
        {"name": "REP Some Loser",   "party_code": "REP", "votes":  "4200"},
        {"name": "Write-in",         "party_code": "NON", "votes":   "100"}
      ]
    },
    { "name": "City Council CITY OF PROVIDENCE",
      "votes_allowed": "3",
      "candidates": [
        {"name": "DEM First Council",  "party_code": "DEM", "votes": "8000"},
        {"name": "DEM Second Council", "party_code": "DEM", "votes": "7500"},
        {"name": "REP Third Council",  "party_code": "REP", "votes": "7000"},
        {"name": "Ind Fourth Loser",   "party_code": "Ind", "votes": "5000"},
        {"name": "Write-in",           "party_code": "NON", "votes":  "200"}
      ]
    },
    { "name": "Town Clerk TOWN OF MIDDLETOWN",
      "votes_allowed": "1",
      "candidates": [
        {"name": "Jane Roe",  "party_code": "NON", "votes": "5000"},
        {"name": "Write-in",  "party_code": "NON", "votes":   "50"}
      ]
    },
    { "name": "School Committee CITY OF NEWPORT",
      "votes_allowed": "2",
      "candidates": [
        {"name": "Pat Doe",   "party_code": "NON", "votes": "5000"},
        {"name": "Sam Roe",   "party_code": "NON", "votes": "4500"},
        {"name": "Sad Loser", "party_code": "NON", "votes": "1000"}
      ]
    },
    { "name": "Mayor TOWN OF ATLANTIS",
      "votes_allowed": "1",
      "candidates": [
        {"name": "Ind Aquaman", "party_code": "Ind", "votes": "5000"}
      ]
    },
    { "name": "1. CONSTITUTIONAL CONVENTION",
      "votes_allowed": "1",
      "candidates": [
        {"name": "Approve", "party_code": "NON", "votes": "3000"},
        {"name": "Reject",  "party_code": "NON", "votes": "5000"}
      ]
    }
  ]
}"""

INLINE_TIER2 = """[
  {"name": "Providence"}, {"name": "Pawtucket"},
  {"name": "Newport"}, {"name": "Middletown"},
  {"name": "Cranston"}, {"name": "Warwick"}
]"""


def self_test() -> int:
    """Happy-path test exercising the full pipeline offline."""
    with tempfile.TemporaryDirectory() as td_str:
        td = Path(td_str)
        db_path = td / "self_test.db"
        f_counties = td / "counties.json"; f_counties.write_text(INLINE_CENSUS_COUNTIES)
        f_cousubs  = td / "cousubs.json";  f_cousubs.write_text(INLINE_CENSUS_COUSUBS)
        f_tier1    = td / "tier1.json";    f_tier1.write_text(INLINE_TIER1)
        f_tier2    = td / "tier2.json";    f_tier2.write_text(INLINE_TIER2)

        conn = open_db(db_path)
        try:
            init_schema(conn)
            seeded = seed_static(conn)
            assert seeded["office_types"] == 6, seeded
            assert seeded["synonyms"] >= 30, seeded   # safety net for off-by-one

            ref_counts = seed_reference_from_census(
                conn,
                counties_from_file=f_counties,
                munis_from_file=f_cousubs,
            )
            assert ref_counts["counties"] == 2, ref_counts
            assert ref_counts["municipalities"] == 4, ref_counts

            # Fuzzy title matcher direct check.
            ot_id, canonical, score = fuzzy_canonical_office_type(
                "Town Council Member", conn,
            )
            assert canonical == "Local Legislative Member", (canonical, score)
            assert score >= FUZZY_TITLE_THRESHOLD

            ot_id, canonical, score = fuzzy_canonical_office_type(
                "Senator in General Assembly", conn,
            )
            # No synonym matches a state-legislator title; reconciler will
            # send it to needs_review.
            assert canonical is None, (canonical, score)

            # Ingest the Tier 1 file via the data-lake path so collection_log
            # gets a real row.
            sos_id = _resolve_source_id(conn, TIER1_NAME)
            _, log_id = ingest_local_file(f_tier1, TIER1_SLUG, conn, sos_id)

            records = parse_ri_sos_winners(f_tier1)
            # See unit test for the fixture's expected 8 winners.
            assert len(records) == 8, len(records)

            counts = reconcile_records(
                records, conn,
                source_id=sos_id, term_start="2025-01-01",
            )
            # Expected reconcile outcome:
            #   loaded (7): Smiley (Mayor PVD), 3x Council (PVD),
            #               Roe (Town Clerk MID), Pat Doe + Sam Roe (School Cmte NPT)
            #   needs_review (1): Aquaman (unknown muni 'Atlantis')
            assert counts["loaded"] == 7, counts
            assert counts["needs_review"] == 1, counts
            assert counts["touched"] == 0, counts
            assert counts["closed_and_opened"] == 0, counts

            # Back-fill the log row's records_found / records_updated.
            update_log_counts(conn, log_id,
                              records_found=len(records),
                              records_updated=counts["loaded"])
            row = conn.execute(
                "SELECT records_found, records_updated FROM collection_log "
                "WHERE log_id = ?", (log_id,),
            ).fetchone()
            assert row[0] == len(records) and row[1] == counts["loaded"], row

            # Idempotency: a second pass should touch every loaded row.
            counts2 = reconcile_records(
                records, conn,
                source_id=sos_id, term_start="2025-01-01",
            )
            assert counts2["touched"] == 7, counts2
            assert counts2["loaded"] == 0, counts2
            assert counts2["closed_and_opened"] == 0, counts2

            # Term lifecycle: a new Providence Mayor in a later cycle should
            # close the old office_holders row and open a new one (single-seat
            # office, different person). This is the Section-3 pseudocode.
            new_mayor = [{
                "muni_name": "Providence", "full_name": "Successor Smiley",
                "office_title": "Mayor", "party": "Democratic",
                "source_url": TIER1_URL,
                "storage_path": str(f_tier1),
            }]
            counts3 = reconcile_records(
                new_mayor, conn,
                source_id=sos_id, term_start="2027-01-01",
            )
            assert counts3["closed_and_opened"] == 1, counts3
            assert counts3["loaded"] == 0, counts3
            mayor_rows = list(conn.execute("""
                SELECT off.full_name, oh.is_current, oh.term_end
                FROM office_holders oh
                JOIN officials off USING(official_id)
                JOIN offices    o   USING(office_id)
                WHERE o.local_title = 'Mayor'
                ORDER BY oh.is_current
            """))
            assert len(mayor_rows) == 2, mayor_rows  # one closed + one current
            closed, current = mayor_rows
            assert closed[0] == "Brett Smiley" and closed[1] == 0 and closed[2] is not None
            assert current[0] == "Successor Smiley" and current[1] == 1

            # needs_review queue persisted the failure each pass.
            review = list(conn.execute(
                "SELECT failure_reason FROM needs_review ORDER BY review_id"
            ))
            # 1 fail per pass x 2 passes (Atlantis on each) = 2 entries.
            assert len(review) == 2, review
            assert all("unknown municipality" in r[0] for r in review)

            # Tier 2 cross-check (loaded munis: Providence, Middletown, Newport).
            xc = cross_check_munis(conn, str(f_tier2))
            assert xc["matched"] == 3 and xc["unmatched"] == 0, xc

            # Schema-traversal sanity check (7 current officeholders after
            # the close-and-open in pass 3: Successor PVD Mayor + 3 Council
            # + Jane Roe MID + Pat Doe NPT + Sam Roe NPT).
            n = conn.execute("""
                SELECT COUNT(*) FROM office_holders oh
                JOIN offices       o  ON o.office_id     = oh.office_id
                JOIN municipalities m ON m.muni_id       = o.muni_id
                JOIN counties      c  ON c.county_fips   = m.county_fips
                JOIN office_types  ot ON ot.office_type_id = o.office_type_id
                JOIN officials     off ON off.official_id = oh.official_id
                WHERE oh.is_current = 1
            """).fetchone()[0]
            assert n == 7, n

            # Post-load sanity checks (Section 4 v1 checks).
            checks = run_post_load_checks(conn)
            by_name = {c["name"]: c for c in checks}
            # 4 munis seeded, 4 with officeholders -> gap rate 0%, passes.
            assert by_name["gap_rate_munis"]["passed"], by_name["gap_rate_munis"]
            # Just-inserted rows -> staleness 0%, passes.
            assert by_name["staleness"]["passed"], by_name["staleness"]

            # collection_log captured the ingest.
            log_count = conn.execute(
                "SELECT COUNT(*) FROM collection_log WHERE status = 'success'"
            ).fetchone()[0]
            assert log_count >= 1, log_count
        finally:
            conn.close()

    print("\nself-test PASSED:")
    print("  - Census fetch loop: 2 counties + 4 munis seeded from fixtures")
    print("  - real RI SOS shape: 7 contests parsed, 8 winners emitted")
    print("    (filtered 1 state-office contest + 1 ballot question)")
    print("  - multi-seat top-N: 3 of 4 City Council candidates loaded")
    print("  - party-prefix stripping: 'DEM Brett Smiley' -> 'Brett Smiley'")
    print("  - fuzzy title matcher: 'Town Council Member' -> Local Legislative Member")
    print("  - 7 records loaded, 1 routed to needs_review (unknown muni 'Atlantis')")
    print("  - idempotent second pass: 7 touched, 0 loaded")
    print("  - term lifecycle: new Mayor closed old + opened new (single-seat)")
    print("  - post-load sanity checks: gap_rate + staleness both pass")
    print("  - Tier 2 cross-check: 3/3 loaded munis matched (PVD/MID/NPT)")
    print("  - schema joins traverse cleanly across 6 tables")
    print("  - collection_log captured the ingest with hash + path")
    return 0


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _term_start_from_election_date(election_date: str | None) -> str:
    """RI municipal terms typically begin Jan 1 after a Nov general election.

    Accepts ISO ('2024-11-05') and the SOS US format ('November 05, 2024' /
    'Nov 5, 2024'). Returns ISO date string. Falls back to today when the
    input is missing or unparseable - the schema requires a non-NULL
    term_start for current rows.
    """
    if not election_date:
        return date.today().isoformat()
    s = str(election_date).strip()
    for fmt in ("%Y-%m-%d", "%B %d, %Y", "%b %d, %Y"):
        try:
            ed = datetime.strptime(s, fmt).date()
            break
        except ValueError:
            continue
    else:
        return date.today().isoformat()
    return (date(ed.year + 1, 1, 1).isoformat()
            if ed.month >= 10 else ed.isoformat())


def cmd_run(args: argparse.Namespace) -> int:
    """End-to-end pipeline against live URLs.

    For RI the script is URL-only by design - file-based ingestion + offline
    reference fetches are useful for testing but not for the production path,
    so they live in self_test() rather than the CLI.
    """
    conn = open_db(args.db_path)
    try:
        init_schema(conn)
        static_counts = seed_static(conn)
        for k, v in static_counts.items():
            print(f"  seeded static  {k:14s} {v}")

        ref_counts = seed_reference_from_census(conn)
        for k, v in ref_counts.items():
            print(f"  fetched ref    {k:14s} {v}")

        sos_id = _resolve_source_id(conn, TIER1_NAME)

        print(f"\nFetching per-muni RI SOS results (base = {args.url})")
        fetched = fetch_ri_sos_all_munis(conn, sos_id, args.url)
        print(f"  fetched {len(fetched)} muni files into the data lake")

        # Per-muni: parse + reconcile + back-fill log counts.
        # Aggregating into a single rollup at the end keeps the per-muni
        # provenance in collection_log while giving the operator one
        # summary line. term_start is derived per file from the parsed
        # election_date - no CLI flag needed.
        rollup = {"loaded": 0, "closed_and_opened": 0,
                  "touched": 0, "needs_review": 0}
        total_records = 0
        for muni, path, log_id in fetched:
            records = parse_ri_sos_winners(
                path, source_url=f"{args.url.rstrip('/')}/{_to_url_slug(muni)}.json",
            )
            term_start = _term_start_from_election_date(
                records[0]["election_date"] if records else None
            )
            counts = reconcile_records(
                records, conn,
                source_id=sos_id, term_start=term_start,
            )
            records_updated = (counts.get("loaded", 0)
                               + counts.get("closed_and_opened", 0)
                               + counts.get("touched", 0))
            update_log_counts(conn, log_id, len(records), records_updated)
            for k in rollup:
                rollup[k] += counts.get(k, 0)
            total_records += len(records)

        print(f"\nReconcile rollup over {len(fetched)} munis "
              f"({total_records} parsed records): {rollup}")

        print(f"\nPost-load sanity checks:")
        for check in run_post_load_checks(conn):
            tag = "PASS" if check["passed"] else "FAIL"
            print(f"  [{tag}] {check['name']}: {check['detail']}")

        if not TIER2_PATH.exists():
            print(
                f"\nTier 2 cross-check skipped: {TIER2_PATH} not found. "
                "Bundle the RI Open Data Portal election-summary XLSX "
                "alongside this script (see README for the source URL).",
                file=sys.stderr,
            )
        else:
            print(f"\nCross-checking against Tier 2: {TIER2_PATH.name}")
            cross_check_munis(conn, str(TIER2_PATH))
    finally:
        conn.close()
    return 0


def cmd_self_test(args: argparse.Namespace) -> int:
    """CLI handler for `self-test`. Delegates to self_test()."""
    return self_test()


def build_parser() -> argparse.ArgumentParser:
    """Build the argparse hierarchy for `run` and `self-test` subcommands."""
    p = argparse.ArgumentParser(
        prog="ri_officials_mvp.py",
        description="RI elected officials data acquisition - single-file MVP.",
    )
    sub = p.add_subparsers(dest="cmd", required=True)

    p_run = sub.add_parser(
        "run",
        help="end-to-end: init + seed + Census fetch + Tier 1 + cross-check.",
    )
    p_run.add_argument("--url", default=RI_SOS_BASE_FETCH_URL,
                       help=("RI SOS per-muni JSON base URL. The script "
                             "appends '/<slug>.json' for every seeded "
                             "muni and fetches each into the data lake. "
                             f"Default: {RI_SOS_BASE_FETCH_URL}"))
    p_run.add_argument("--db-path", default=DEFAULT_DB_PATH)
    p_run.set_defaults(func=cmd_run)

    p_st = sub.add_parser("self-test", help="Run the offline happy-path test.")
    p_st.set_defaults(func=cmd_self_test)
    return p


def main(argv: list[str] | None = None) -> int:
    """Module entry point. Parses argv and dispatches to the subcommand handler."""
    args = build_parser().parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
